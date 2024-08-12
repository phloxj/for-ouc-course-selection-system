import openpyxl
import re
import sys
import os

# 输入数据
path = r"C:\Users\qwe_r\Documents\选课"  # 文件夹路径
year = '2024'  # 学年
term = '1'  # 学期 夏-0 秋-1 春-2

# 构建路径
read_name = "\\xuanke" + year + term + ".xlsx"
read_path = path + read_name  # 读取路径
write_name = "\\preprocess" + year + term + ".xlsx"
write_path = path + write_name  # 写入路径
# print(write_path)

# 创建写入文件
if os.path.exists(write_path):
    # 如果文件已存在，则报错
    print(f"文件 {write_path} 已存在，无法创建新文件。")
    sys.exit()
else:
    # 创建
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    sheet1.title = 'Specialty'
    sheet2 = workbook.create_sheet(title='Common')
    sheet3 = workbook.create_sheet(title='PublicBasic')
    workbook.save(write_path)
    # print("文件已成功创建。")


# 汉字转数值
def shuz(zhi):
    weekStr = "一二三四五六日"
    shu = weekStr.find(zhi) + 1
    return shu


def write_data(k):  # 输入sheet_name，写入rows
    # 指定文件路径 打开工作表
    file_pat = write_path  # 写入路径
    workboo = openpyxl.load_workbook(file_pat)
    sheet_nam = k
    shee = workboo[sheet_nam]
    # 数据写入
    for row in rows:
        shee.append(row)
    workboo.save(file_pat)
    return


def transposition(L):
    return list(map(list, zip(*L)))


# 打开工作表
file_path = read_path  # 打开路径
workbook = openpyxl.load_workbook(file_path)

kcfw = ['Specialty', 'Common', 'PublicBasic']  # 课程范围
for kc in kcfw:
    sheet_name = kc
    sheet = workbook[sheet_name]
    # 上课时间处理，time
    obj = re.compile(r'周 (?P<zhou>.*?)[(](?P<staj>\d+)-(?P<endj>\d+)节')
    time = []
    column_values = [cell.value for cell in sheet['M']]
    for value in column_values:
        if value is not None:
            ans = obj.finditer(value)

            for i in ans:
                time += [[shuz(i.group('zhou')), i.group('staj'), i.group('endj')]]
        else:
            time += [['']]

    # 课程处理，course
    obj = re.compile(r'\[\d+](?P<num>.*)')
    course = []
    column_values = [cell.value for cell in sheet['B']]
    for value in column_values:
        if value is not None:
            ans = obj.finditer(value)
            for i in ans:
                course += [[i.group('num')]]
        else:
            course += [['']]
    # print(course)

    number = []
    column_values = [cell.value for cell in sheet['A']]
    for value in column_values:
        if value is not None:
            number += [[value]]
        else:
            number += [['']]
    del number[0]

    rows = transposition(transposition(number) + transposition(course) + transposition(time))
    write_data(kc)
print("预处理成功")
