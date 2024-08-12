import os
import sys
import time
import base64
import openpyxl
import requests
from aip import AipOcr
from hashlib import md5
from bs4 import BeautifulSoup

# 输入数据
APP_ID = ''  # 在百度官网的应用列表中查看APP_ID
API_KEY = ''  # 在百度官网的应用列表中查看API_KEY
SECRET_KEY = ''  # 在百度官网的应用列表中查看SECRET_KEY

username = ''  # 用户名
password = ''  # 密码

year = '2023'  # 学年
term = '1'  # 学期 夏-0 秋-1 春-2

path = r"C:\Users\qwe_r\Documents\选课"  # 文件夹路径
# 开始时间
start_time = time.time()

# 创建表格
file_name = "\\xuanke" + year + term + ".xlsx"
file_path = path + file_name  # 写入路径
if os.path.exists(file_path):
    # 如果文件已存在，则报错
    print(f"文件 {file_path} 已存在，无法创建新文件。")
    sys.exit()
else:
    # 创建
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    sheet1.title = 'Specialty'
    sheet2 = workbook.create_sheet(title='Common')
    sheet3 = workbook.create_sheet(title='PublicBasic')
    workbook.save(file_path)

# 一.模拟登录

url = 'http://jwgl.ouc.edu.cn/cas/logon.action'
data = {}
headers = {
    'Host': 'jwgl.ouc.edu.cn',
    'Origin': 'http://jwgl.ouc.edu.cn',
    'Referer': 'http://jwgl.ouc.edu.cn/cas/login.action',
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36',
    'X-Requested-With': 'XMLHttpRequest',
    'Upgrade-Insecure-Requests': '1',
}


# 百度OCR
def randnumber_ocr(image):
    client = AipOcr(APP_ID, API_KEY, SECRET_KEY)
    text = client.basicAccurate(image)
    if text['words_result_num'] == 1:
        return text['words_result'][0]['words'].strip()
    else:
        return '识别错误'


# 验证码
randnumber = ''
session = requests.Session()
while len(randnumber) != 4:
    r = session.get('http://jwgl.ouc.edu.cn/cas/genValidateCode', headers=headers)
    randnumber = randnumber_ocr(r.content).replace(' ', '')

data['randnumber'] = randnumber
p_username = '_u' + randnumber
p_password = '_p' + randnumber

# username加密 base64
_sessionid = session.cookies.get_dict()['JSESSIONID']
username = base64.b64encode(username.encode('utf-8') + b';;' + str(_sessionid).encode('utf-8'))
data[p_username] = username

# paseword加密 MD5哈希
password = md5(password.encode('utf-8')).hexdigest()
randnumber = md5(randnumber.lower().encode('utf-8')).hexdigest()
password = md5((password + randnumber).encode('utf-8')).hexdigest()
data[p_password] = password

r = session.post(url, data=data, headers=headers)
print(r.text)


# 二.转为表格

# 解析r,输出rows
def get_row():
    soup = BeautifulSoup(r.text, 'html.parser')
    table = soup.find('table')
    ro = []
    for row in table.find_all('tr'):
        cols = row.find_all('td')
        cols = [element.text.strip() for element in cols]
        ro.append(cols)
    return ro


# 将数据写入指定工作表
def write_data(k):
    # 指定文件路径 打开工作表
    workbook0 = openpyxl.load_workbook(file_path)
    sheet_name = k
    sheet = workbook0[sheet_name]
    # 数据写入
    for row in rows:
        sheet.append(row)
    workbook0.save(file_path)
    return


url = 'http://jwgl.ouc.edu.cn/taglib/DataTable.jsp'
formdata = {
    'initQry': '0',
    'xktype': '2',  # 选课类型
    'xh': username,  # 学号
    'xn': year,  # 学年
    'xq': term,  # 学期 夏-0 秋-1 春-2
    'nj': '2023',  # 年级
    'zydm': '0006',  # 专业代码
    'items': '',
    'xnxq': year + '-' + term,  # 学年 - 学期
    'sel_nj': '2023',  # 选择_年级
    'sel_zydm': '0006',  # 选择_专业代码
    'sel_schoolarea': '2',  # 校区代码
    'sel_cddwdm': '',
    'sel_kc': '',
    'kcmc': '',
    'tableId': '6146',
}
headers['Referer'] = 'http://jwgl.ouc.edu.cn/taglib/DataTable.jsp?tableId=6146'

# 开始循环
kcfw = ['Specialty', 'Common', 'PublicBasic']  # 课程范围
for kc in kcfw:
    print("处理项目：", kc)
    url = 'http://jwgl.ouc.edu.cn/taglib/DataTable.jsp'
    formdata['kcfw'] = kc
    r = session.post(url, data=formdata, headers=headers)
    rows = get_row()
    across = len(rows)
    write_data(kc)
    print(f"第 1 页，共 {across} 行数据")

    page = 2
    while across >= 101:
        url = 'http://jwgl.ouc.edu.cn/taglib/DataTable.jsp?currPageCount=' + str(page)
        r = session.post(url, data=formdata, headers=headers)
        rows = get_row()
        across = len(rows)
        del rows[0]
        write_data(kc)
        print(f"第 {page} 页，共 {across} 行数据")
        page += 1

print('完成')

# 结束时间
end_time = time.time()

# 计算时间差
elapsed_time = end_time - start_time
print(f"操作耗时: {elapsed_time:.2f} 秒")
