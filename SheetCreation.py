import sys
import os
import openpyxl

# 指定文件路径
file_path = r"C:\Users\qwe_r\Documents\选课\kon.xlsx"

if os.path.exists(file_path):
    # 如果文件已存在，则报错
    print(f"文件 {file_path} 已存在，无法创建新文件。")
    sys.exit()
else:
    # 创建
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    sheet1.title = 'Specialty'
    sheet2 = workbook.create_sheet(title='Common')
    sheet3 = workbook.create_sheet(title='PublicBasic')
    workbook.save(file_path)
    # print("文件已成功创建。")
